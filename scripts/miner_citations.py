import argparse
import os
import sys
import time
import urllib.parse
import requests
from bs4 import BeautifulSoup
import openpyxl

# --- 1. Carrega a chave do arquivo .env na raiz do projeto (se existir) ---
ENV_PATH = ".env"
if os.path.exists(ENV_PATH):
    with open(ENV_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line.startswith("SCRAPER_API_KEY="):
                os.environ["SCRAPER_API_KEY"] = line.split("=", 1)[1].strip().strip('"').strip("'")

SCRAPER_API_KEY = os.getenv("SCRAPER_API_KEY", "")
EXCEL_PATH = os.path.join("data", "JSON Generator", "dataWASHES-data.xlsx")

def get_apa_citation_from_scholar(title, retries=3):
    """Busca o artigo no Google Scholar via ScraperAPI usando render=true."""
    if not SCRAPER_API_KEY:
        print("   ⚠️ 'SCRAPER_API_KEY' não configurada no .env.")
        return None

    # URL do Google Scholar
    scholar_url = f"https://scholar.google.com/scholar?q={urllib.parse.quote(title)}&hl=pt-BR"
    
    # Parâmetros oficiais recomendados pela ScraperAPI para o Google Scholar
    payload = {
        'api_key': SCRAPER_API_KEY,
        'url': scholar_url,
        'country_code': 'us',
        'render': 'true'  # 👈 ATIVA O NAVEGADOR PARA BURLAR O CAPTCHA DO GOOGLE SCHOLAR
    }

    for attempt in range(1, retries + 1):
        try:
            response = requests.get('https://api.scraperapi.com', params=payload, timeout=60)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, "html.parser")
                first_result = soup.find("div", class_="gs_r")
                if not first_result:
                    return "#"  # Não encontrado no Scholar

                cited_by_link = soup.find("a", string=lambda t: t and ("Citado por" in t or "Cited by" in t))
                if not cited_by_link:
                    return "#"  # Encontrado, mas com 0 citações

                return cited_by_link.text.strip()
            
            elif response.status_code in [401, 403]:
                print("   ❌ Chave ScraperAPI inválida ou limite de créditos atingido.")
                return None

        except requests.exceptions.RequestException as e:
            print(f"   ⚠️ Tentativa {attempt}/{retries} aguardando resposta da ScraperAPI... Retentando em 3s.")
            time.sleep(3)

    return None

def run_miner(target_year=None, force=False, limit=None, dry_run=False):
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Planilha não encontrada em: {EXCEL_PATH}")
        return

    if not SCRAPER_API_KEY:
        print("❌ 'SCRAPER_API_KEY' não foi encontrada no .env!")
        return

    print(f"📖 Lendo planilha em: {EXCEL_PATH}")
    if dry_run:
        print("🔍 Modo DRY-RUN: nenhuma alteração será gravada na planilha.")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_year = headers.index("Year") + 1 if "Year" in headers else 2
    col_title = headers.index("Paper's title") + 1 if "Paper's title" in headers else 3
    col_cites = headers.index("Citações") + 1 if "Citações" in headers else 15

    updated_count = 0
    checked_count = 0

    for row in range(2, ws.max_row + 1):
        if limit is not None and checked_count >= limit:
            print(f"⏹️ Limite de {limit} artigos analisados atingido. Encerrando.")
            break

        year_val = ws.cell(row=row, column=col_year).value
        title = ws.cell(row=row, column=col_title).value
        current_cite = ws.cell(row=row, column=col_cites).value

        if target_year and str(year_val) != str(target_year):
            continue

        # Se force=True, testa mesmo que a célula já tenha texto
        is_candidate = force or (current_cite is None or str(current_cite).strip() in ["", "#"])

        if title and is_candidate:
            checked_count += 1
            print(f"\n🔍 [{checked_count}] Testando no Scholar (Ano {year_val}): '{title[:50]}...'")
            print(f"   📌 Valor atual na planilha: {str(current_cite)[:60]}...")

            citation = get_apa_citation_from_scholar(title)

            if citation is not None:
                log_value = '0' if citation == '#' else citation
                print(f"   [INFO] Artigo {checked_count}: {log_value} citações encontradas")
                if citation != "#":
                    action = "[DRY-RUN] seria atualizado para:" if dry_run else "saved to:"
                    print(f"   ℹ️ Citações capturadas → {action} {citation}")
                    if not dry_run:
                        ws.cell(row=row, column=col_cites).value = citation
                    updated_count += 1
                else:
                    print("   ℹ️ Confirmado: 0 citações no Google Scholar.")
            else:
                print("   ⏩ Pulado devido a erro de resposta do proxy.")

            time.sleep(1)

    if limit is not None and checked_count > 0:
        print(f"⏹️ {checked_count} artigos analisados (limite: {limit}).")

    if updated_count > 0:
        if dry_run:
            print(f"\n[Dry-run] {updated_count} citações seriam atualizadas — planilha intacta.")
        else:
            wb.save(EXCEL_PATH)
            print(f"\n🎉 Sucesso! {updated_count} citações foram salvas em '{EXCEL_PATH}'.")
    else:
        print(f"\n✨ Nenhuma citação nova ({checked_count} artigos analisados).")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Minera citações do Google Scholar")
    parser.add_argument("--year", type=int, help="Filtrar por ano específico (ex: --year 2016)")
    parser.add_argument("--force", action="store_true", help="Força re-checagem mesmo de artigos que já possuem citações")
    parser.add_argument("--limit", type=int, default=None, help="Analisa no máximo N artigos (ex: --limit 3)")
    parser.add_argument("--dry-run", action="store_true", help="Não grava na planilha: apenas reporta o que seria atualizado")
    args = parser.parse_args()

    run_miner(target_year=args.year, force=args.force, limit=args.limit, dry_run=args.dry_run)